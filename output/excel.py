"""Excel 결과 출력"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

_CYCLE_FILL = {
    "RED":    PatternFill("solid", fgColor="FFB3B3"),
    "YELLOW": PatternFill("solid", fgColor="FFF3B3"),
    "GREEN":  PatternFill("solid", fgColor="B3FFB3"),
}

_HEADERS = [
    "등수", "단지명", "지역", "면적(㎡)", "평당가(만원)",
    "현재호가(만원)", "최근실거래가(만원)", "실거래일",
    "전고점(만원)", "전고점시점", "호가회복률(%)", "실거래회복률(%)",
    "상태", "매물링크",
]

# column indices (1-based) for conditional formatting
_COL_ASK_RATE   = 11  # 호가회복률(%)
_COL_TRADE_RATE = 12  # 실거래회복률(%)
_COL_CYCLE      = 13  # 상태


def _row_values(apt: dict) -> list:
    return [
        apt.get("rank", ""),
        apt.get("complex_name", ""),
        apt.get("area_name", ""),
        apt.get("area", ""),
        apt.get("price_per_pyeong", ""),
        apt.get("ask_price", ""),
        apt.get("latest_trade_price", ""),
        apt.get("latest_trade_date", ""),
        apt.get("peak_price", ""),
        apt.get("peak_date", ""),
        apt.get("ask_rate", ""),
        apt.get("trade_rate", ""),
        apt.get("cycle", ""),
        apt.get("url", ""),
    ]


def _write_sheet(ws, rows: list[dict]) -> None:
    ws.append(_HEADERS)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for apt in rows:
        ws.append(_row_values(apt))

    # cycle color on 상태 column
    for row_idx in range(2, ws.max_row + 1):
        cycle = ws.cell(row_idx, _COL_CYCLE).value
        fill = _CYCLE_FILL.get(cycle)
        if fill:
            ws.cell(row_idx, _COL_CYCLE).fill = fill

    # color scale on 호가회복률 / 실거래회복률 (green=low, red=high)
    last_row = ws.max_row
    if last_row >= 2:
        for col in (_COL_ASK_RATE, _COL_TRADE_RATE):
            col_letter = get_column_letter(col)
            cell_range = f"{col_letter}2:{col_letter}{last_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="num", start_value=0,   start_color="B3FFB3",
                    mid_type="num",   mid_value=80,    mid_color="FFF3B3",
                    end_type="num",   end_value=100,   end_color="FFB3B3",
                ),
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _write_summary_sheet(ws, all_apts: list[dict]) -> None:
    headers = [
        "지역", "분석단지수", "평균평당가", "평균호가회복률",
        "평균실거래회복률", "최저회복률단지", "최고회복률단지",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    by_area: dict[str, list[dict]] = {}
    for apt in all_apts:
        by_area.setdefault(apt.get("area_name", ""), []).append(apt)

    for area_name, apts in by_area.items():
        n = len(apts)
        avg_ppp  = round(sum(a.get("price_per_pyeong", 0) for a in apts) / n)
        avg_ask  = round(sum(a.get("ask_rate",  0) for a in apts) / n, 1)
        avg_trd  = round(sum(a.get("trade_rate", 0) for a in apts) / n, 1)
        min_apt  = min(apts, key=lambda a: a.get("ask_rate", 0))
        max_apt  = max(apts, key=lambda a: a.get("ask_rate", 0))
        ws.append([
            area_name, n, avg_ppp, avg_ask, avg_trd,
            min_apt.get("complex_name", ""),
            max_apt.get("complex_name", ""),
        ])

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def export_to_excel(
    all_apts: list[dict],
    filtered_apts: list[dict],
    filepath: str,
) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "전체 랭킹"
    _write_sheet(ws1, all_apts)

    ws2 = wb.create_sheet("필터 통과 매물")
    _write_sheet(ws2, filtered_apts)

    ws3 = wb.create_sheet("지역별 요약")
    _write_summary_sheet(ws3, all_apts)

    wb.save(filepath)
